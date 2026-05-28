from openpyxl import load_workbook

from qa_jira.excel import write_bugsheet
from qa_jira.models import BugInEpic, Epic


def _bug(idx: int) -> BugInEpic:
    return BugInEpic(
        key=f"PROJ-{idx}",
        summary=f"Bug {idx}",
        status="Open",
        priority="P1",
        assignee="Alice",
        reporter="Bob",
        created="2026-05-01",
        description="see prod logs",
        url=f"https://x.atlassian.net/browse/PROJ-{idx}",
        environment="Production",
    )


def test_write_bugsheet(tmp_path):
    epic = Epic(key="PROJ-1", summary="Login work")
    out = write_bugsheet(bugs=[_bug(2), _bug(3)], epic=epic, output_dir=tmp_path)

    assert out.exists()
    assert out.name.startswith("bugsheet-PROJ-1-")
    assert out.suffix == ".xlsx"

    wb = load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[0] == "Bug ID"
    assert headers[4] == "JIRA ID"
    assert ws.cell(row=2, column=1).value == "BUG_ID_1"
    assert ws.cell(row=2, column=2).value == "Bug"
    formula = ws.cell(row=2, column=5).value
    assert formula.startswith("=HYPERLINK(")
    assert "PROJ-2" in formula
    assert ws.freeze_panes == "A2"


def test_write_bugsheet_handles_empty_assignee(tmp_path):
    bug = _bug(9)
    bug.assignee = ""
    out = write_bugsheet(bugs=[bug], epic=Epic(key="PROJ-1", summary="x"), output_dir=tmp_path)
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=2, column=11).value == "Unassigned"
