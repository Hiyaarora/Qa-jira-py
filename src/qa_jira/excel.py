from __future__ import annotations

from datetime import date
from pathlib import Path

from dateutil.parser import parse as parse_date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from qa_jira.models import BugInEpic, Epic

HEADERS = [
    "Bug ID",
    "Bug Type",
    "Reported By",
    "Reporting Date",
    "JIRA ID",
    "Title",
    "Current Status",
    "Environment",
    "Priority",
    "RCA",
    "Assignee",
    "Remarks",
]

HEADER_FILL = PatternFill("solid", fgColor="1565C0")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ROW_FILL_LIGHT = PatternFill("solid", fgColor="E3F2FD")
CENTER = Alignment(horizontal="center", vertical="center")


def _format_date(s: str) -> str:
    if not s:
        return ""
    try:
        return parse_date(s).strftime("%d-%b-%Y")
    except (ValueError, TypeError):
        return s


def write_bugsheet(*, bugs: list[BugInEpic], epic: Epic, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    filename = f"bugsheet-{epic.key}-{today}.xlsx"
    out_path = output_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = f"{epic.key} — Bug Sheet"[:31]

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for row_idx, bug in enumerate(bugs, start=2):
        bug_id_label = f"BUG_ID_{row_idx - 1}"
        ws.cell(row=row_idx, column=1, value=bug_id_label)
        ws.cell(row=row_idx, column=2, value="Bug")
        ws.cell(row=row_idx, column=3, value=bug.reporter or "")
        ws.cell(row=row_idx, column=4, value=_format_date(bug.created))
        ws.cell(row=row_idx, column=5, value=f'=HYPERLINK("{bug.url}","{bug.key}")')
        ws.cell(row=row_idx, column=6, value=bug.summary or "")
        ws.cell(row=row_idx, column=7, value=bug.status or "")
        ws.cell(row=row_idx, column=8, value=bug.environment or "UAT")
        ws.cell(row=row_idx, column=9, value=bug.priority or "")
        ws.cell(row=row_idx, column=10, value="")  # RCA
        ws.cell(row=row_idx, column=11, value=bug.assignee or "Unassigned")
        ws.cell(row=row_idx, column=12, value="")  # Remarks

        if (row_idx - 2) % 2 == 1:
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ROW_FILL_LIGHT

    ws.freeze_panes = "A2"

    for col_idx in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, ws.max_row + 1)
            ),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

    wb.save(out_path)
    return out_path
